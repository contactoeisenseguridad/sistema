import os
import random
import string
import requests
import openpyxl  # 👈 LIBRERÍA NUEVA PARA LEER EXCEL
from datetime import timedelta

from django.conf import settings
from django.contrib import admin, messages
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.core.mail import EmailMessage, send_mail
from django.db.models import OuterRef, Subquery, Sum
from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.utils.html import format_html
from django import forms

from .models import (
    Alumno, Inscripcion, Auditoria, Aviso, Pago, Cuota, GastoOperacional,
    PerfilUsuario, Modulo, SesionClase, Asistencia, PlanillaSPD, PlantillaDocumento
)

# ==========================================
# 🛡️ SISTEMA DE AUDITORÍA CENTRALIZADO
# ==========================================

def registrar_auditoria(request, obj, accion):
    """
    Función maestra para registrar movimientos. 
    Usa un try/except para evitar que errores de texto tumben el servidor.
    """
    try:
        # Intentamos obtener una representación legible del objeto
        identificador = str(obj)
    except:
        identificador = f"ID: {obj.id}"

    try:
        Auditoria.objects.create(
            usuario=request.user.username,
            accion=accion,
            modelo=obj._meta.verbose_name.upper(),
            objeto_id=obj.id or 0,
            descripcion=f"El usuario {request.user.username} realizó {accion} en: {identificador}"
        )
    except Exception as e:
        # Si la auditoría falla, que no detenga la operación principal del sistema
        print(f"Error registrando auditoría: {e}")

@admin.register(Auditoria)
class AuditoriaAdmin(admin.ModelAdmin):
    list_display = ('fecha', 'usuario', 'accion', 'modelo', 'descripcion')
    list_filter = ('accion', 'modelo', 'usuario', 'fecha')
    search_fields = ('usuario', 'descripcion', 'modelo')
    
    # Bloqueo total de edición manual
    readonly_fields = ('usuario', 'accion', 'modelo', 'objeto_id', 'descripcion', 'fecha')

    def has_add_permission(self, request): 
        return False
    
    def has_delete_permission(self, request, obj=None): 
        return False

    def has_change_permission(self, request, obj=None):
        # Permite ver el detalle pero no guardar cambios
        return False

# ==========================================
# 1. INTEGRACIÓN MOODLE Y FUNCIONES ÚTILES
# ==========================================

def enviar_a_moodle(inscripcion):
    MOODLE_URL = "https://virtual.otecuno.cl/webservice/rest/server.php"
    MOODLE_TOKEN = "401791078af1d393dce611bd34c9549e" 
    
    CURSOS_MOODLE = {'FGS': 73, 'CCTV': 71, 'SSPP': 72}
    
    try:
        sigla_curso = None
        grupo_str = str(inscripcion.grupo).upper() if inscripcion.grupo else ""
        curso_str = str(inscripcion.curso).upper() if inscripcion.curso else ""
        
        if "FGS" in grupo_str: sigla_curso = 'FGS'
        elif "CCTV" in curso_str: sigla_curso = 'CCTV'
        elif "SSPP" in curso_str: sigla_curso = 'SSPP'
        
        if not sigla_curso or sigla_curso not in CURSOS_MOODLE:
            return False, f"No reconocí la sigla del curso en: '{curso_str}' o '{grupo_str}'"

        curso_id_moodle = CURSOS_MOODLE[sigla_curso]
        
        correo_limpio = str(inscripcion.alumno.correo).strip().lower()
        nombres_limpios = str(inscripcion.alumno.nombres).strip()
        apellidos_limpios = str(inscripcion.alumno.apellidos).strip()
        rut_original = str(inscripcion.alumno.rut).strip()
        
        rut_solo_numeros = rut_original.split('-')[0].replace('.', '').strip()
        username = rut_solo_numeros.lower()
        password = username 
        
        params_buscar = {
            'wstoken': MOODLE_TOKEN, 'wsfunction': 'core_user_get_users_by_field',
            'moodlewsrestformat': 'json', 'field': 'username', 'values[0]': username
        }
        busqueda = requests.post(MOODLE_URL, data=params_buscar).json()
        
        moodle_user_id = None
        if isinstance(busqueda, list) and len(busqueda) > 0:
            moodle_user_id = busqueda[0]['id']
        else:
            params_crear = {
                'wstoken': MOODLE_TOKEN, 'wsfunction': 'core_user_create_users',
                'moodlewsrestformat': 'json', 'users[0][username]': username,
                'users[0][password]': password, 'users[0][firstname]': nombres_limpios,
                'users[0][lastname]': apellidos_limpios, 'users[0][email]': correo_limpio,
                'users[0][idnumber]': rut_original, 
            }
            creacion = requests.post(MOODLE_URL, data=params_crear).json()
            if isinstance(creacion, dict) and 'exception' in creacion:
                return False, f"🔴 Falló la creación en Moodle | Detalle: {creacion.get('message', '')}"
            if isinstance(creacion, list) and len(creacion) > 0 and 'id' in creacion[0]:
                moodle_user_id = creacion[0]['id']
            else:
                return False, "No se obtuvo ID al crear usuario."
        
        params_matricular = {
            'wstoken': MOODLE_TOKEN, 'wsfunction': 'enrol_manual_enrol_users',
            'moodlewsrestformat': 'json', 'enrolments[0][roleid]': 5, 
            'enrolments[0][userid]': moodle_user_id, 'enrolments[0][courseid]': curso_id_moodle
        }
        matricula = requests.post(MOODLE_URL, data=params_matricular).json()
        if isinstance(matricula, dict) and 'exception' in matricula:
            return False, f"Error al matricular: {matricula.get('message')}"
            
    except Exception as error_general:
        return False, f"Falló: {str(error_general)}"
        
    try:
        cuerpo_bienvenida = f"Estimad@ {nombres_limpios}:\n\nSu acceso a la plataforma virtual (http://virtual.otecuno.cl) ha sido habilitado.\nUsuario y Contraseña: Su RUN sin puntos (Ej: 12345678-9).\n\nSaludos,\nOTEC UNO"
        correo_bienvenida = EmailMessage(
            subject='Acceso a Plataforma Virtual - OTEC Uno', body=cuerpo_bienvenida,
            from_email=settings.DEFAULT_FROM_EMAIL, to=[correo_limpio],
        )
        ruta_pdf = os.path.join(settings.BASE_DIR, 'manual_aula_virtual.pdf')
        if os.path.exists(ruta_pdf):
            correo_bienvenida.attach_file(ruta_pdf)
        correo_bienvenida.send(fail_silently=False)
    except Exception as error_correo:
        return False, "Matriculado en Moodle, pero falló el correo."
        
    return True, "Matriculado con éxito en Moodle y correo enviado."

def formato_clp(valor): 
    return f"$ {valor:,.0f}.-".replace(",", ".") if valor else "$ 0.-"

def validar_rut_chileno(rut):
    rut = rut.upper().replace(".", "").replace("-", "").strip()
    if len(rut) < 2 or not rut[:-1].isdigit(): return False
    suma, multiplo = 0, 2
    for numero in reversed(rut[:-1]):
        suma += int(numero) * multiplo
        multiplo = 2 if multiplo == 7 else multiplo + 1
    dv_calculado = 11 - (suma % 11)
    dv_calculado = "0" if dv_calculado == 11 else "K" if dv_calculado == 10 else str(dv_calculado)
    return rut[-1] == dv_calculado

def formatear_rut(rut):
    rut = rut.upper().replace(".", "").replace("-", "").strip()
    return f"{rut[:-1]}-{rut[-1]}"


# ==========================================
# 2. INLINES Y PANELES BÁSICOS
# ==========================================

class InscripcionInline(admin.TabularInline):
    model = Inscripcion
    extra = 1

class PagoInline(admin.TabularInline):
    model = Pago
    extra = 0
    fields = ('inscripcion', 'metodo_pago', 'monto_total', 'cantidad_cuotas', 'fecha_pago', 'enlace_detalle')
    readonly_fields = ('enlace_detalle',)
    
    def enlace_detalle(self, obj):
        if obj.pk: 
            return format_html('<a class="button" href="/admin/alumnos/pago/{}/change/" target="_blank">💸 Administrar Cuotas</a>', obj.pk)
        return "Guarde el alumno para generar pagos"

class CuotaInline(admin.TabularInline):
    model = Cuota
    extra = 1
    fields = ('numero_cuota', 'monto', 'fecha_vencimiento', 'estado')

    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser and obj: 
            return ('numero_cuota', 'monto', 'fecha_vencimiento', 'fecha_pago')
        return super().get_readonly_fields(request, obj)

    def has_delete_permission(self, request, obj=None): 
        return request.user.is_superuser


# ==========================================
# 3. PANELES DE ADMINISTRACIÓN (ADMINS)
# ==========================================

admin.site.register(Aviso)

@admin.register(Alumno)
class AlumnoAdmin(admin.ModelAdmin):
    list_display = (
        'apellidos', 'nombres', 'rut', 'estado_rut', 'grupo_actual', 'estado_correo'
    )
    search_fields = ('nombres', 'apellidos', 'rut', 'correo')
    list_filter = ('correo_confirmado', 'rut_confirmado')
    list_per_page = 100
    ordering = ('apellidos', 'nombres')
    
    inlines = [InscripcionInline]
    
    fields = (
        'nombres', 'apellidos', 'rut', 'rut_confirmado', 'direccion', 'comuna', 
        'correo', 'telefono', 'fecha_registro', 'boton_enviar_codigo', 
        'correo_confirmado', 'codigo_ingresado', 'boton_validar_codigo', 'ficha_alumno'
    )
    readonly_fields = (
        'fecha_registro', 'rut_confirmado', 'correo_confirmado', 
        'boton_enviar_codigo', 'boton_validar_codigo', 'ficha_alumno'
    )

    # --- FUNCIÓN ÚNICA DE GUARDADO (FUSIONADA) ---
    def save_model(self, request, obj, form, change):
        # A. Lógica de Auditoría
        accion_auditoria = "EDICIÓN" if change else "CREACIÓN"
        
        # B. Lógica de RUT
        rut_original = obj.rut
        if obj.rut:
            rut_limpio = obj.rut.upper().replace(".", "").replace("-", "").strip()
            if validar_rut_chileno(rut_limpio):
                obj.rut = formatear_rut(rut_limpio)
                obj.rut_confirmado = True
            else:
                obj.rut_confirmado = False
                messages.error(request, f"El RUT ingresado no es válido: {rut_original}")
                Auditoria.objects.create(
                    usuario=request.user.username, 
                    accion="RUT INVALIDO", 
                    modelo="ALUMNO", 
                    objeto_id=obj.id or 0, 
                    descripcion=f"RUT inválido ingresado: {rut_original}"
                )

        # C. Lógica de Validación de Código de Correo
        if "_validar_codigo" in request.POST:
            codigo_ingresado = str(obj.codigo_ingresado).strip() if obj.codigo_ingresado else ""
            codigo_real = str(obj.codigo_confirmacion).strip() if obj.codigo_confirmacion else ""
            
            if obj.intentos_codigo >= 3: 
                messages.error(request, "Demasiados intentos. Debe enviar un nuevo código.")
            elif not obj.fecha_codigo: 
                messages.error(request, "Debe enviar primero un código al correo.")
            elif timezone.now() > (obj.fecha_codigo + timedelta(minutes=10)): 
                messages.error(request, "El código expiró. Debe enviar un nuevo código.")
            elif codigo_ingresado and codigo_ingresado == codigo_real:
                obj.correo_confirmado = True
                obj.codigo_ingresado = None
                obj.intentos_codigo = 0
                messages.success(request, "Correo confirmado correctamente ✅")
                registrar_auditoria(request, obj, "CONFIRMAR CORREO")
            else:
                obj.intentos_codigo += 1
                messages.error(request, f"Código incorrecto ❌ Intento {obj.intentos_codigo} de 3.")

        # D. Guardado final y Auditoría General
        super().save_model(request, obj, form, change)
        registrar_auditoria(request, obj, accion_auditoria)

    # --- FUNCIONES DE ESTADO Y BOTONES ---
    def estado_rut(self, obj): 
        return "✅ Válido" if obj.rut_confirmado else "❌ Pendiente/Inválido"
    estado_rut.short_description = "RUT"

    def estado_correo(self, obj): 
        return "✅ Confirmado" if obj.correo_confirmado else "❌ Pendiente"
    estado_correo.short_description = "Correo"

    def boton_enviar_codigo(self, obj):
        if obj.id: 
            return format_html(
                '<a class="button" href="/admin/alumnos/alumno/enviar-codigo/{}/">📧 Enviar código al correo</a>', 
                obj.id
            )
        return "Primero debe guardar el alumno."
    boton_enviar_codigo.short_description = "Confirmación de correo"

    def boton_validar_codigo(self, obj):
        if obj.id: 
            return format_html(
                '<button type="submit" name="_validar_codigo" class="button">{}</button>', 
                "✅ Validar código"
            )
        return "Primero debe guardar el alumno."
    boton_validar_codigo.short_description = "Validar correo"

    def ficha_alumno(self, obj):
        if obj.id: 
            return format_html(
                '<a class="button" href="/alumno/{}/" target="_blank">🔎 Ver ficha / Generar contrato</a>', 
                obj.id
            )
        return "Primero debe guardar el alumno."
    ficha_alumno.short_description = "Ficha del alumno"

    # --- GESTIÓN DE URLS Y QUERIES ---
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('enviar-codigo/<int:alumno_id>/', self.admin_site.admin_view(self.enviar_codigo))
        ]
        return custom_urls + urls

    def enviar_codigo(self, request, alumno_id):
        alumno = Alumno.objects.get(id=alumno_id)
        if not alumno.correo:
            messages.error(request, "El alumno no tiene correo registrado.")
            return HttpResponseRedirect(f"/admin/alumnos/alumno/{alumno.id}/change/")
            
        alumno.codigo_confirmacion = str(random.randint(1000, 9999))
        alumno.fecha_codigo = timezone.now()
        alumno.intentos_codigo = 0
        alumno.codigo_ingresado = None
        alumno.correo_confirmado = False
        alumno.save()
        
        email = EmailMessage(
            subject='Código de confirmación de correo',
            body=(f'Estimado/a {alumno.nombres},\n\n'
                  f'Su código de confirmación es: {alumno.codigo_confirmacion}\n\n'
                  f'Este código tendrá una vigencia de 10 minutos.\n\n'
                  f'OTEC Uno EIRL\nPreocupados por tu futuro'),
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[alumno.correo],
        )
        try:
            email.send(fail_silently=False)
            registrar_auditoria(request, alumno, "ENVIAR CODIGO CORREO")
            messages.success(request, f"Código enviado correctamente a {alumno.correo}.")
        except Exception as e:
            messages.error(request, f"No se pudo enviar el correo: {e}")
            
        return HttpResponseRedirect(f"/admin/alumnos/alumno/{alumno.id}/change/")

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        ultima_inscripcion = Inscripcion.objects.filter(alumno=OuterRef('pk')).order_by('-fecha_inicio')
        return qs.annotate(grupo_orden=Subquery(ultima_inscripcion.values('grupo')[:1]))

    def grupo_actual(self, obj): 
        return obj.grupo_orden or '-'
    grupo_actual.short_description = 'Grupo'
    grupo_actual.admin_order_field = 'grupo_orden'

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        if search_term:
            alumnos_por_grupo = Inscripcion.objects.filter(
                grupo__icontains=search_term
            ).values_list('alumno_id', flat=True)
            queryset |= self.model.objects.filter(id__in=alumnos_por_grupo)
        return queryset, use_distinct

    def has_change_permission(self, request, obj=None): 
        return super().has_change_permission(request, obj)
        
    def has_delete_permission(self, request, obj=None): 
        return request.user.is_superuser

    def response_change(self, request, obj):
        if "_validar_codigo" in request.POST: 
            return HttpResponseRedirect(f"/admin/alumnos/alumno/{obj.id}/change/")
        return super().response_change(request, obj)

    def response_add(self, request, obj, post_url_continue=None):
        return HttpResponseRedirect(f"/admin/alumnos/alumno/{obj.id}/")

@admin.register(Inscripcion)
class InscripcionAdmin(admin.ModelAdmin):
    list_display = ('alumno', 'curso', 'grupo', 'fecha_inicio', 'estado')
    search_fields = ('alumno__nombres', 'alumno__apellidos', 'alumno__rut', 'grupo')
    list_filter = ('grupo', 'curso')
    actions = ['matricular_masivamente_moodle']

    def matricular_masivamente_moodle(self, request, queryset):
        exitos, errores = 0, 0
        for i in queryset:
            if not i.alumno.correo:
                messages.error(request, f"❌ {i.alumno} no tiene correo.")
                continue
            exito, msj = enviar_a_moodle(i)
            if exito: exitos += 1
            else: messages.error(request, f"❌ Error en {i.alumno}: {msj}")
        if exitos > 0: messages.success(request, f"✅ {exitos} alumnos matriculados.")
    matricular_masivamente_moodle.short_description = "🎓 Matricular seleccionados en Moodle"


@admin.register(Pago)
class PagoAdmin(admin.ModelAdmin):
    # 1. Configuración del Listado (Unificada)
    list_display = ('alumno', 'formato_monto', 'metodo_pago')
    search_fields = ('alumno__nombres', 'alumno__apellidos', 'alumno__rut')
    list_filter = ('metodo_pago',)
    autocomplete_fields = ['alumno']
    
    # 2. Configuración del Formulario
    fields = ('alumno', 'inscripcion', 'metodo_pago', 'monto_total', 'cantidad_cuotas', 'observaciones')
    inlines = [CuotaInline]

    # --- LÓGICA DE FILTRADO DINÁMICO CORREGIDA ---
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "inscripcion":
            parent_obj_id = request.resolver_match.kwargs.get('object_id')
            if parent_obj_id:
                try:
                    pago = Pago.objects.get(pk=parent_obj_id)
                    if pago.alumno:
                        # IMPORTANTE: Inscripcion con "I" mayúscula
                        kwargs["queryset"] = Inscripcion.objects.filter(alumno=pago.alumno)
                except (Pago.DoesNotExist, NameError):
                    # Si hay error o el pago no existe, muestra un queryset vacío o el default
                    pass
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # --- FORMATO DE DINERO $XXX.XXX ---
    def formato_monto(self, obj):
        if obj.monto_total:
            return f"${obj.monto_total:,.0f}".replace(",", ".")
        return "$0"
    formato_monto.short_description = 'Monto Total'

    def save_model(self, request, obj, form, change):
        accion = "EDICIÓN DE PAGO" if change else "CREACIÓN DE PAGO"
        super().save_model(request, obj, form, change)
        registrar_auditoria(request, obj, accion)
    
@admin.register(Cuota)
class CuotaAdmin(admin.ModelAdmin):
    # Punto 6: Listado de deudores con formato moneda y nombres claros
    list_display = ('alumno_nombre', 'numero_cuota', 'formato_monto_cuota', 'fecha_vencimiento', 'estado')
    list_filter = ('estado', 'fecha_vencimiento')
    
    # IMPORTANTE: search_fields permite que el buscador de arriba encuentre por RUT o Nombre
    search_fields = ('pago__alumno__nombres', 'pago__alumno__apellidos', 'pago__alumno__rut')
    
    # Optimización: Carga el pago y el alumno de una sola vez para que sea rápido
    list_select_related = ('pago__alumno',)

    # --- FORMATO DE DINERO $XXX.XXX ---
    def formato_monto_cuota(self, obj):
        if obj.monto:
            return f"${obj.monto:,.0f}".replace(",", ".")
        return "$0"
    formato_monto_cuota.short_description = 'Monto Cuota'

    # --- NOMBRE DEL ALUMNO (PROTEGIDO) ---
    def alumno_nombre(self, obj):
        try:
            # Verificamos que exista el pago y el alumno antes de mostrarlo
            if obj.pago and obj.pago.alumno:
                return f"{obj.pago.alumno}"
        except Exception:
            pass
        return "Sin Alumno"
    alumno_nombre.short_description = "Alumno"


# ==========================================
# 4. NUEVOS PANELES: PROGRAMACIÓN Y EXCEL
# ==========================================

@admin.register(PerfilUsuario)
class PerfilUsuarioAdmin(admin.ModelAdmin):
    list_display = ('user', 'rut', 'rol', 'bloqueado', 'intentos_fallidos')
    list_filter = ('rol', 'bloqueado')
    search_fields = ('user__username', 'user__first_name')
    actions = ['desbloquear_usuarios']

    def desbloquear_usuarios(self, request, queryset):
        queryset.update(bloqueado=False, intentos_fallidos=0)
        messages.success(request, "✅ Usuarios desbloqueados exitosamente.")
    desbloquear_usuarios.short_description = "🔓 Desbloquear Relatores seleccionados"


@admin.register(Modulo)
class ModuloAdmin(admin.ModelAdmin):
    list_display = ('nombre',)
    search_fields = ('nombre',)


@admin.register(SesionClase)
class SesionClaseAdmin(admin.ModelAdmin):
    # Organiza por línea de tiempo arriba
    date_hierarchy = 'fecha'
    
    # Columnas que pertenecen estrictamente a una Sesión de Clase (¡Incluyendo el Relator!)
    list_display = ('fecha', 'bloque_horario', 'modulo', 'grupo', 'relator')
    
    # Filtros laterales correctos para las clases
    list_filter = ('grupo', 'modulo', 'relator')
    
    # Buscador de clases por grupo o nombre del módulo
    search_fields = ('grupo', 'modulo__nombre')
    
    ordering = ('-fecha', 'bloque_horario')


@admin.register(Asistencia)
class AsistenciaAdmin(admin.ModelAdmin):
    # Columnas calculadas para ver la asistencia de forma clara y fácil
    list_display = ('get_alumno_nombre', 'get_alumno_rut', 'get_grupo_codigo', 'get_sesion_info', 'presente')
    
    # Filtros laterales para agrupar las asistencias por Grupo y por Estado
    list_filter = ('sesion_clase__grupo', 'presente')
    
    # Buscador inteligente de asistencias
    search_fields = ('alumno__apellidos', 'alumno__nombres', 'alumno__rut', 'sesion_clase__grupo')
    
    # Optimización para las llaves foráneas correspondientes a Asistencia
    raw_id_fields = ('alumno', 'sesion_clase')

    # Métodos seguros para mostrar los datos relacionales sin romper la página
    def get_alumno_nombre(self, obj):
        try:
            return f"{obj.alumno.apellidos}, {obj.alumno.nombres}".upper()
        except Exception:
            return str(obj.alumno)
    get_alumno_nombre.short_description = 'Alumno'

    def get_alumno_rut(self, obj):
        try:
            return obj.alumno.rut
        except Exception:
            return "-"
    get_alumno_rut.short_description = 'RUT'

    def get_grupo_codigo(self, obj):
        try:
            return obj.sesion_clase.grupo
        except Exception:
            return "-"
    get_grupo_codigo.short_description = 'Grupo'

    def get_sesion_info(self, obj):
        try:
            return f"{obj.sesion_clase.fecha.strftime('%d/%m/%Y')} | {obj.sesion_clase.modulo.nombre}"
        except Exception:
            return str(obj.sesion_clase)
    get_sesion_info.short_description = 'Sesión de Clase'

@admin.register(PlanillaSPD)
class PlanillaSPDAdmin(admin.ModelAdmin):
    list_display = ('grupo', 'fecha_subida', 'procesado')
    actions = ['procesar_excel_spd_action']

    def procesar_excel_spd_action(self, request, queryset):
        import openpyxl
        exitos = 0
        for planilla in queryset:
            try:
                wb = openpyxl.load_workbook(planilla.archivo_excel.path, data_only=True)
                pestaña = "CONTENIDOS DEL CURSO"
                
                if pestaña in wb.sheetnames:
                    ws = wb[pestaña]
                    # Limpiamos sesiones previas del grupo antes de cargar las nuevas
                    SesionClase.objects.filter(grupo=planilla.grupo).delete()
                    
                    # RANGO FIJO: Fila 11 hasta la 33
                    for r in range(11, 34): 
                        # B=2 (Fecha), C=3 (Horario), D=4 (Módulo), E=5 (Relator)
                        fecha_val = ws.cell(row=r, column=2).value
                        horario_val = ws.cell(row=r, column=3).value
                        modulo_val = ws.cell(row=r, column=4).value
                        relator_val = ws.cell(row=r, column=5).value

                        # Si la celda de fecha está vacía, saltamos esa fila del rango
                        if not fecha_val:
                            continue

                        try:
                            # 1. Buscamos o creamos el Módulo (Columna D)
                            mod_obj, _ = Modulo.objects.get_or_create(
                                nombre=str(modulo_val).strip()
                            )
                            
                            # 2. IDENTIFICAR AL RELATOR (Columna E)
                            relator_obj = None
                            if relator_val:
                                nombre_busqueda = str(relator_val).strip()
                                # Intentamos buscar al relator. Asumiendo que tu modelo Relator tiene campo 'apellidos' o 'nombres'
                                # Si tu modelo busca por el campo 'nombre_completo' o similar, ajusta el filtro abajo.
                                # Buscaremos si el texto del Excel coincide o está contenido en sus datos:
                                from .models import Relator # Asegúrate de que el modelo esté importado
                                relator_obj = Relator.objects.filter(
                                    apellidos__icontains=nombre_busqueda
                                ).first() or Relator.objects.filter(
                                    nombres__icontains=nombre_busqueda
                                ).first()
                            
                            # 3. Creamos la Sesión de Clase incluyendo el campo 'relator'
                            SesionClase.objects.create(
                                grupo=planilla.grupo,
                                fecha=fecha_val,
                                bloque_horario=str(horario_val).strip() if horario_val else "Sin Horario",
                                modulo=mod_obj,
                                relator=relator_obj # 👈 ¡CON ESTO QUEDA IDENTIFICADO Y ASOCIADO!
                            )
                            exitos += 1
                        except Exception as row_error:
                            # print(f"Error fila {r}: {str(row_error)}") # Descomenta si necesitas debuguear en local
                            continue

                planilla.procesado = (exitos > 0)
                planilla.save()
                
            except Exception as e:
                messages.error(request, f"Error en {planilla.grupo}: {str(e)}")

        if exitos > 0:
            messages.success(request, f"✅ ¡TRABAJO TERMINADO! Se cargaron {exitos} sesiones vinculando sus módulos y relatores.")
        else:
            messages.warning(request, "No se encontraron datos válidos en el rango B11:F33.")

# ==========================================
# 5. VISTA DEL RELATOR (ASISTENCIA)
# ==========================================

def vista_pasar_asistencia(request):
    if 'relator_auth' not in request.session:
        if request.method == "POST" and "auth_relator" in request.POST:
            user_input = request.POST.get('usuario')
            pass_input = request.POST.get('clave')
            try:
                perfil = PerfilUsuario.objects.get(user__username=user_input, rol='RELATOR')
                if perfil.bloqueado:
                    return render(request, 'error_asistencia.html', {'error': 'USUARIO BLOQUEADO.'})
                user = authenticate(username=user_input, password=pass_input)
                if user is not None:
                    perfil.intentos_fallidos = 0
                    perfil.save()
                    request.session['relator_auth'] = user.id
                else:
                    perfil.intentos_fallidos += 1
                    if perfil.intentos_fallidos >= 2: perfil.bloqueado = True
                    perfil.save()
                    return render(request, 'login_relator.html', {'error': f'Clave incorrecta.'})
            except PerfilUsuario.DoesNotExist:
                return render(request, 'login_relator.html', {'error': 'Usuario no autorizado.'})

    if request.method == "POST" and "seleccionar_grupo" in request.POST:
        grupo_id = request.POST.get('grupo')
        modulo_id = request.POST.get('modulo')
        alumnos = Alumno.objects.filter(inscripcion__grupo=grupo_id)
        return render(request, 'lista_asistencia.html', {'alumnos': alumnos, 'grupo': grupo_id, 'modulo': modulo_id})

    if request.method == "POST" and "enviar_asistencia" in request.POST:
        grupo_id = request.POST.get('grupo')
        modulo_id = request.POST.get('modulo')
        alumnos_presentes_ids = request.POST.getlist('alumnos_presentes')
        sesion = SesionClase.objects.filter(grupo=grupo_id, modulo_id=modulo_id, fecha=timezone.now().date()).first()
        if not sesion:
            return render(request, 'error_asistencia.html', {'error': 'No hay clases hoy.'})
        for alumno in Alumno.objects.filter(inscripcion__grupo=grupo_id):
            es_presente = str(alumno.id) in alumnos_presentes_ids
            Asistencia.objects.update_or_create(alumno=alumno, sesion=sesion, defaults={'presente': es_presente})
        return render(request, 'exito_asistencia.html')

    return render(request, 'login_relator.html')


# ==========================================
# 6. PORTAL DE ASISTENCIA
# ==========================================

def portal_asistencia(request):
    modulos = Modulo.objects.all()
    
    if request.method == "POST":
        modulo_id = request.POST.get('modulo')
        # Convertimos grupo a MAYÚSCULAS automáticamente
        grupo_input = request.POST.get('grupo').strip().upper()
        
        # Cambia esa línea por esta:
        modulo = get_object_or_404(Modulo, id=modulo_id)
        
        # 1. Buscar la sesión de hoy para ese grupo y módulo
        hoy = timezone.now().date()
        sesion = SesionClase.objects.filter(
            grupo=grupo_input, 
            modulo=modulo, 
            fecha=hoy
        ).first()

        if not sesion:
            messages.error(request, f"No existe una clase programada para hoy ({hoy}) en el grupo {grupo_input} y módulo {modulo.nombre}.")
            return render(request, 'portal_asistencia.html', {'modulos': modulos})

        # 2. Si el relator ya envió la lista de alumnos
        if 'guardar_asistencia' in request.POST:
            alumnos_inscritos = Alumno.objects.filter(inscripcion__grupo=grupo_input)
            
            for alumno in alumnos_inscritos:
                # El checkbox solo envía valor si está marcado (Presente)
                estado_check = request.POST.get(f'asistencia_{alumno.id}')
                es_presente = True if estado_check == 'presente' else False
                
                # Guardar o actualizar asistencia
                asistencia_obj, created = Asistencia.objects.update_or_create(
                    alumno=alumno,
                    sesion=sesion,
                    defaults={'presente': es_presente}
                )

                # 3. Enviar Correo Electrónico
                estado_texto = "PRESENTE" if es_presente else "AUSENTE"
                try:
                    send_mail(
                        subject=f'Registro de Asistencia - {modulo.nombre}',
                        message=f'Estimado(a) {alumno.nombres} {alumno.apellidos}:\n\n'
                                f'Le informamos que su asistencia para el día de hoy {hoy} '
                                f'en el módulo "{modulo.nombre}" ha sido registrada como: {estado_texto}.\n\n'
                                f'Cualquier discrepancia debe ser indicada al correo: contacto@otecuno.cl\n\n'
                                f'Atentamente,\nOTEC UNO.',
                        from_email='contacto@otecuno.cl',
                        recipient_list=[alumno.correo],
                        fail_silently=True,
                    )
                except:
                    pass
            
            messages.success(request, f"Asistencia del grupo {grupo_input} guardada y correos enviados.")
            return render(request, 'exito.html')

        # 4. Cargar listado de alumnos para ese grupo
        alumnos = Alumno.objects.filter(inscripcion__grupo=grupo_input).order_by('apellidos')
        return render(request, 'portal_asistencia.html', {
            'alumnos': alumnos, 
            'grupo': grupo_input, 
            'modulo': modulo,
            'sesion': sesion
        })

    return render(request, 'portal_asistencia.html', {'modulos': modulos})

# ==========================================
# REPOSITORIO DE DOCUMENTOS
# ==========================================

@admin.register(PlantillaDocumento)
class PlantillaDocumentoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'codigo_referencia', 'fecha_creacion')
    prepopulated_fields = {'codigo_referencia': ('nombre',)}

    def save_model(self, request, obj, form, change):
        accion = "EDICIÓN DE PLANTILLA" if change else "CREACIÓN DE PLANTILLA"
        super().save_model(request, obj, form, change)
        registrar_auditoria(request, obj, accion)


# ==========================================
# CONFIGURACIÓN VISUAL DEL PANEL
# ==========================================
admin.site.site_header = "Sistema Operativo - OTEC Uno"
admin.site.site_title = "Administración - OTEC Uno"
admin.site.index_title = "Panel de Control General"