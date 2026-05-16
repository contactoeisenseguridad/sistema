from datetime import datetime
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.core.mail import send_mail, EmailMessage
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.db.models import Sum
import os
from .models import Alumno, Inscripcion, PlantillaDocumento, SesionClase, Asistencia, Modulo
from django.template import Template, Context


import openpyxl

# Importación de todos tus modelos
from .models import (
    Alumno, Inscripcion, Auditoria, Aviso, 
    PerfilUsuario, Modulo, SesionClase, Asistencia
)

def limpiar(valor):
    if valor is None:
        return ''
    return str(valor).strip()

# ==========================================
# 🔵 FUNCIONES ORIGINALES (MANTENIDAS)
# ==========================================

@login_required
def inicio(request):
    aviso = Aviso.objects.last()
    return render(request, 'inicio.html', {'aviso': aviso})

@login_required
@permission_required('alumnos.view_alumno', raise_exception=True)
def buscar_alumno(request):
    query = request.GET.get('q')
    alumnos = []
    if query:
        alumnos = Alumno.objects.filter(
            apellidos__icontains=query
        ) | Alumno.objects.filter(
            rut__icontains=query
        ) | Alumno.objects.filter(
            nombres__icontains=query
        ) | Alumno.objects.filter(
            inscripciones__grupo__icontains=query
        )
        alumnos = alumnos.distinct()

        Auditoria.objects.create(
            usuario=request.user.username,
            accion="BUSCAR",
            modelo="Alumno",
            objeto_id=0,
            descripcion=f"Búsqueda realizada con término: {query}"
        )
    return render(request, 'buscar.html', {'alumnos': alumnos, 'query': query})

@login_required
@permission_required('alumnos.view_alumno', raise_exception=True)
def detalle_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, id=alumno_id)
    inscripciones = alumno.inscripciones.all()
    Auditoria.objects.create(
        usuario=request.user.username,
        accion="VER FICHA",
        modelo="Alumno",
        objeto_id=alumno.id,
        descripcion=f"Se visualizó la ficha del alumno {alumno.nombres} {alumno.apellidos}"
    )
    return render(request, 'detalle.html', {'alumno': alumno, 'inscripciones': inscripciones})

@login_required
@permission_required('alumnos.view_alumno', raise_exception=True)
def generar_pdf(request, alumno_id):
    alumno = get_object_or_404(Alumno, id=alumno_id)
    Auditoria.objects.create(
        usuario=request.user.username,
        accion="VER CONTRATO HTML",
        modelo="Alumno",
        objeto_id=alumno.id,
        descripcion=f"Se visualizó el contrato HTML del alumno {alumno.nombres} {alumno.apellidos}"
    )
    return render(request, 'alumnos/contrato_pdf.html', {
        'alumno': alumno,
        'fecha': datetime.now(),
    })

@login_required
@permission_required('alumnos.view_alumno', raise_exception=True)
def formulario_exportar_excel(request):
    return render(request, 'exportar_excel.html')

@login_required
@permission_required('alumnos.view_alumno', raise_exception=True)
def exportar_excel_grupo(request):
    grupo = request.GET.get('grupo')
    inscripciones = Inscripcion.objects.filter(
        grupo__iexact=grupo
    ).select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    ws.append(["RUN", "APELLIDOS", "NOMBRES", "EMAIL", "DIRECCION", "COMUNA", "TELEFONO"])

    for ins in inscripciones:
        a = ins.alumno
        ws.append([a.rut, a.apellidos, a.nombres, a.correo, a.direccion, a.comuna, a.telefono])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="grupo_{grupo}.xlsx"'
    wb.save(response)

    Auditoria.objects.create(
        usuario=request.user.username,
        accion="EXPORTAR EXCEL",
        modelo="Inscripcion",
        objeto_id=0,
        descripcion=f"Exportación de alumnos del grupo {grupo}"
    )
    return response

@login_required
@permission_required('alumnos.add_alumno', raise_exception=True)
def carga_masiva_alumnos(request):
    resultado = None
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        creados, actualizados, inscripciones_creadas = 0, 0, 0
        errores = []
        encabezados = [limpiar(cell.value).lower() for cell in ws[1]]

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            datos = dict(zip(encabezados, row))
            try:
                rut = limpiar(datos.get('rut')).upper()
                if not rut:
                    errores.append(f"Fila {idx}: RUT vacío")
                    continue
                alumno, creado = Alumno.objects.update_or_create(
                    rut=rut,
                    defaults={
                        'apellidos': limpiar(datos.get('apellidos')),
                        'nombres': limpiar(datos.get('nombres')),
                        'correo': limpiar(datos.get('correo')),
                        'direccion': limpiar(datos.get('direccion')),
                        'comuna': limpiar(datos.get('comuna')),
                        'telefono': limpiar(datos.get('telefono')),
                    }
                )
                if creado: creados += 1
                else: actualizados += 1

                curso, grupo = limpiar(datos.get('curso')), limpiar(datos.get('grupo'))
                if curso and grupo:
                    _, ins_creada = Inscripcion.objects.get_or_create(
                        alumno=alumno, curso=curso, grupo=grupo,
                        defaults={
                            'fecha_inicio': datos.get('fecha_inicio') or None,
                            'fecha_fin': datos.get('fecha_fin') or None,
                            'estado': limpiar(datos.get('estado')),
                            'observaciones': limpiar(datos.get('observaciones')),
                        }
                    )
                    if ins_creada: inscripciones_creadas += 1
            except Exception as e:
                errores.append(f"Fila {idx}: {str(e)}")

        Auditoria.objects.create(
            usuario=request.user.username,
            accion="CARGA MASIVA",
            modelo="Alumno",
            objeto_id=0,
            descripcion=f"Carga masiva: {creados} creados, {actualizados} actualizados."
        )
        resultado = {'creados': creados, 'actualizados': actualizados, 'inscripciones_creadas': inscripciones_creadas, 'errores': errores}

    return render(request, 'carga_masiva.html', {'resultado': resultado})

# ==========================================
# 🌐 PORTAL DE ASISTENCIA Y FISCALIZACIÓN
# ==========================================

@login_required
def portal_asistencia(request):
    es_admin = request.user.is_superuser
    perfil = getattr(request.user, 'perfilusuario', None)
    
    # Si no tiene perfil configurado y no es administrador, bloqueamos el acceso
    if not perfil and not es_admin:
        return render(request, 'error_acceso.html', {'error': 'Usuario sin Perfil configurado.'})

    modulos = Modulo.objects.all()
    context = {
        'modulos': modulos, 
        'perfil': perfil, 
        'hoy': timezone.now().date(),
        'es_admin': es_admin
    }

    if request.method == "POST":
        # 1. Validación de PIN (El superusuario se salta esta medida por comodidad)
        if not es_admin:
            pin_ingresado = request.POST.get('pin_seguridad')
            if pin_ingresado != perfil.pin:
                messages.error(request, "PIN de seguridad incorrecto. Verifique sus credenciales.")
                return render(request, 'portal_asistencia.html', context)

        # 2. Rescate de datos del formulario
        modulo_id = request.POST.get('modulo')
        grupo_input = request.POST.get('grupo', '').strip().upper()
        
        if not modulo_id or not grupo_input:
            messages.warning(request, "Debe seleccionar un módulo e ingresar un grupo.")
            return render(request, 'portal_asistencia.html', context)

        modulo_obj = get_object_or_404(Modulo, id=modulo_id)
        
        # 3. Validar Sesión de Clase (Si es admin, trae la última de ese grupo/módulo. Si es relator, solo hoy)
        filtro_sesion = {'grupo': grupo_input, 'modulo': modulo_obj}
        if not es_admin:
            filtro_sesion['fecha'] = timezone.now().date()

        # Buscamos la sesión correspondiente
        sesion = SesionClase.objects.filter(**filtro_sesion).order_by('-fecha').first()

        # Si no existe en absoluto
        if not sesion:
            if es_admin:
                messages.error(request, f"No existe ninguna sesión registrada en el calendario para el grupo {grupo_input} en el módulo {modulo_obj.nombre}.")
            else:
                messages.error(request, f"No hay clase programada hoy para {grupo_input} en {modulo_obj.nombre}.")
            return render(request, 'portal_asistencia.html', context)

        # 🔍 DETECTOR INTELIGENTE DE CAMPOS (Evita el Error 500 de base de datos)
        # Averiguamos dinámicamente si el modelo Asistencia usa el nombre 'sesion' o 'sesion_clase'
        campo_sesion = 'sesion'
        if not hasattr(Asistencia, 'sesion') and hasattr(Asistencia, 'sesion_clase'):
            campo_sesion = 'sesion_clase'

        # --- CASO A: GUARDAR ASISTENCIA (Botón Guardar) ---
        if 'btn_guardar' in request.POST:
            if not es_admin and (perfil and perfil.rol != 'RELATOR'):
                messages.error(request, "Acceso denegado: Solo personal autorizado puede registrar asistencia.")
                return redirect('portal_asistencia')

            alumnos_inscritos = Alumno.objects.filter(inscripciones__grupo=grupo_input).distinct()
            correos_enviados = 0
            
            for alumno in alumnos_inscritos:
                estado = request.POST.get(f'asistencia_{alumno.id}')
                es_presente = (estado == 'presente')

                # Armamos los filtros dinámicos según el campo real detectado
                filtro_asistencia = {'alumno': alumno, campo_sesion: sesion}
                asistencia_previa = Asistencia.objects.filter(**filtro_asistencia).first()
                
                debe_notificar = False
                if not asistencia_previa:
                    debe_notificar = True
                elif asistencia_previa.presente != es_presente:
                    debe_notificar = True

                # Guardamos o actualizamos de forma segura en la BD
                valores_update = {'presente': es_presente}
                Asistencia.objects.update_or_create(
                    alumno=alumno, 
                    **{campo_sesion: sesion},
                    defaults=valores_update
                )

                if debe_notificar and alumno.correo:
                    asunto = f"Registro Asistencia - {modulo_obj.nombre}"
                    estado_txt = "PRESENTE" if es_presente else "AUSENTE"
                    msg = (f"Estimado(a) {alumno.nombres} {alumno.apellidos}:\n\n"
                           f"Le informamos que su asistencia para la sesión del día {sesion.fecha} "
                           f"en el módulo {modulo_obj.nombre} ha sido registrada como: {estado_txt}.\n\n"
                           f"Si existe algún error, contacte a la coordinación o escriba a contacto@otecuno.cl")
                    
                    send_mail(asunto, msg, settings.DEFAULT_FROM_EMAIL, [alumno.correo], fail_silently=True)
                    correos_enviados += 1

            messages.success(request, f"Asistencia de {grupo_input} actualizada. Se enviaron {correos_enviados} notificaciones por email.")
            return redirect(f'/asistencia/?grupo={grupo_input}&modulo={modulo_id}')

        # --- CASO B: CARGAR LISTADO EN PANTALLA ---
        alumnos = Alumno.objects.filter(inscripciones__grupo=grupo_input).distinct().order_by('apellidos')
        
        # Filtramos las asistencias actuales usando el campo dinámico detectado
        filtro_actuales = {campo_sesion: sesion}
        asistencias_actuales = Asistencia.objects.filter(**filtro_actuales)
        
        mapa_asistencia = {a.alumno_id: a.presente for a in asistencias_actuales}
        for a in alumnos:
            a.asistencia_actual = mapa_asistencia.get(a.id, None)
        
        context.update({
            'alumnos': alumnos,
            'grupo_seleccionado': grupo_input,
            'modulo_seleccionado': modulo_obj,
            'sesion': sesion,
            'mapa_asistencia': mapa_asistencia,
            'es_fiscalizador': (perfil and perfil.rol == 'FISCALIZADOR')
        })

    return render(request, 'portal_asistencia.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def buzon_masivo(request):
    grupos = Inscripcion.objects.values_list('grupo', flat=True).distinct().exclude(grupo__isnull=True).exclude(grupo__exact='')

    if request.method == 'POST':
        grupo_seleccionado = request.POST.get('grupo')
        tipo_correo = request.POST.get('tipo_correo')
        horario_curso = request.POST.get('horario_curso', '') # 👈 Capturamos lo que pegues en la caja

        if not grupo_seleccionado or not tipo_correo:
            messages.error(request, "Debe seleccionar un grupo y un tipo de correo.")
            return redirect('buzon_masivo')

        alumnos_del_grupo = Alumno.objects.filter(inscripciones__grupo=grupo_seleccionado).distinct()
        exitos = 0
        errores = 0

        for alumno in alumnos_del_grupo:
            if not alumno.correo:
                errores += 1
                continue

            # 📩 1. CORREO DE BIENVENIDA CON ADJUNTOS
            if tipo_correo == 'bienvenida':
                # Calculamos el nombre de usuario (RUT sin puntos ni dígito verificador)
                rut_usuario = alumno.rut.split('-')[0].replace('.', '').strip()
                
                asunto = "Bienvenido a la plataforma virtual - OTEC UNO"
                cuerpo = (
                    f"Estimado(a) {alumno.nombres} {alumno.apellidos},\n\n"
                    f"Junto con saludar, le damos la bienvenida a nuestra plataforma virtual de capacitación.\n"
                    f"Le informamos que ya puede acceder a la plataforma a través del siguiente enlace:\n"
                    f"http://virtual.otecuno.cl\n\n"
                    f"Su nombre de usuario corresponde a su RUT sin dígito verificador ni puntos.\n"
                    f"Ejemplo:\n"
                    f"RUT: 12.345.678-9\n"
                    f"Usuario: 12345678\n\n"
                    f"Sus datos exactos de acceso son:\n"
                    f"Usuario: {rut_usuario}\n"
                    f"Contraseña: {rut_usuario} (Contraseña por defecto)\n\n"
                    f"En caso de cualquier duda o inconveniente de acceso, puede comunicarse con nosotros al correo: contacto@otecuno.cl\n\n"
                    f"Además, se adjuntan los siguientes documentos:\n"
                    f"- Manual paso a paso para acceder a la plataforma virtual.\n"
                    f"- Documento con todos los links de las clases correspondientes a su curso.\n\n"
                    f"Tus clases son los días y en horarios:\n"
                    f"{horario_curso}\n\n"
                    f"Agradecemos su preferencia y le deseamos mucho éxito en su proceso de capacitación.\n\n"
                    f"Atentamente,\n"
                    f"OTEC UNO E.I.R.L.\n"
                    f"Alameda Libertador Bernardo O’Higgins 1112 Oficina 802, Santiago\n"
                    f"www.otecuno.cl"
                )

                # Creamos el correo avanzado para poder meterle adjuntos
                email = EmailMessage(
                    subject=asunto,
                    body=cuerpo,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[alumno.correo],
                )

                # Rutas a los archivos (Deben estar en la carpeta raíz de tu proyecto, junto a manage.py)
                ruta_manual = os.path.join(settings.BASE_DIR, 'manual_aula_virtual.pdf')
                ruta_links = os.path.join(settings.BASE_DIR, 'links_clases.pdf')

                if os.path.exists(ruta_manual):
                    email.attach_file(ruta_manual)
                if os.path.exists(ruta_links):
                    email.attach_file(ruta_links)

                try:
                    email.send(fail_silently=False)
                    exitos += 1
                except Exception:
                    errores += 1

            # 📜 2. CORREO DE COMPROBANTE NORMAL
            elif tipo_correo == 'comprobante':
                inscripcion = alumno.inscripciones.filter(grupo=grupo_seleccionado).last()
                asunto = f"Comprobante de Matrícula - {inscripcion.curso if inscripcion else 'Curso'}"
                cuerpo = (f"COMPROBANTE DE MATRÍCULA OFICIAL\n"
                          f"--------------------------------------------------\n"
                          f"Nombre: {alumno.nombres} {alumno.apellidos}\n"
                          f"RUT: {alumno.rut}\n"
                          f"Grupo: {grupo_seleccionado}\n"
                          f"Fecha: {timezone.now().strftime('%d/%m/%Y')}\n"
                          f"--------------------------------------------------\n"
                          f"Ficha digital: https://sistema.otecuno.app/alumno/{alumno.id}/\n")

                try:
                    send_mail(asunto, cuerpo, settings.DEFAULT_FROM_EMAIL, [alumno.correo], fail_silently=False)
                    exitos += 1
                except Exception:
                    errores += 1

        if exitos > 0:
            messages.success(request, f"✅ ¡Éxito! Se enviaron {exitos} correos de '{tipo_correo}' al grupo {grupo_seleccionado}.")
        if errores > 0:
            messages.warning(request, f"⚠️ {errores} alumnos no recibieron el correo (sin email registrado o error de servidor).")

        return redirect('buzon_masivo')

    return render(request, 'buzon_masivo.html', {'grupos': grupos})

def visor_repositorio_documentos(request):

    plantillas = PlantillaDocumento.objects.all()
    todos_los_alumnos = Alumno.objects.all().order_by('apellidos')
    documentos_finales = []

    if request.method == "POST":
        plantilla_id = request.POST.get('plantilla')
        grupo_input = request.POST.get('grupo', '').strip().upper()
        alumnos_ids = request.POST.getlist('alumnos_ids')

        plantilla = get_object_or_404(PlantillaDocumento, id=plantilla_id)
        ids_a_procesar = set()

        if grupo_input:
            inscritos = Inscripcion.objects.filter(
                grupo__iexact=grupo_input
            ).values_list('alumno_id', flat=True)

            for aid in inscritos:
                ids_a_procesar.add(aid)

        for aid in alumnos_ids:
            if aid:
                ids_a_procesar.add(int(aid))

        # Generamos documentos
        for alu_id in ids_a_procesar:
            alumno = Alumno.objects.filter(id=alu_id).first()

            if alumno:
                # Buscar la inscripción más reciente
                inscripcion = Inscripcion.objects.filter(
                    alumno=alumno
                ).order_by('-fecha_inicio').first()

                # Usar grupo manual o grupo real
                grupo_alumno = grupo_input or (
                    inscripcion.grupo if inscripcion else ''
                )

                t = Template(plantilla.cuerpo_html)

                c = Context({
                    'nombres': alumno.nombres,
                    'apellidos': alumno.apellidos,
                    'rut': alumno.rut,
                    'grupo': grupo_alumno,
                    'fecha_hoy': timezone.now().strftime('%d/%m/%Y'),
                })

                documentos_finales.append(t.render(c))

        # SI HAY DOCUMENTOS, VAMOS AL VISOR
        if documentos_finales:
            return render(
                request,
                'repositorio/visor_impresion.html',
                {'documentos': documentos_finales}
            )

    return render(request, 'repositorio/repositorio-documentos.html', {
        'plantillas': plantillas,
        'todos_los_alumnos': todos_los_alumnos
    })

def tomar_asistencia(request, sesion_id):
    from django.shortcuts import render, get_object_or_404
    from django.utils import timezone
    from django.contrib import messages
    from .models import SesionClase, Asistencia
    
    sesion = get_object_or_404(SesionClase, id=sesion_id)
    hoy = timezone.now().date()
    
    # TRUCO MAESTRO: Si es superusuario, ignora cualquier bloqueo de fecha
    es_admin = request.user.is_superuser
    
    if not es_admin:
        # Aquí van tus reglas normales para relatores
        if sesion.fecha > hoy:
            return render(request, 'asistencia/error.html', {'mensaje': "No puedes pasar asistencia en una sesión futura."})
        elif sesion.fecha < hoy:
            # Si bloqueas las pasadas, ponlo aquí
            pass

    if request.method == "POST":
        # ... (Tu lógica actual para guardar los presentes/ausentes) ...
        messages.success(request, "Asistencia actualizada correctamente.")
        
    # Obtener alumnos inscritos para mostrar el listado
    # ...